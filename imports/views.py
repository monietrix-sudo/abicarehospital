"""
AbiCare - Excel Import / Export Views
=======================================
KEY RULE: Hospital numbers are ALWAYS auto-generated.
They are NEVER taken from the Excel file.
This prevents PostgreSQL primary key conflicts entirely.
Old/previous hospital numbers go in the 'Old Hospital Number' column only.
"""

import io
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from django.utils import timezone

from .models import ImportSession, ImportError
from apps.audit_logs.utils import log_action


# ── Column name in Excel → model field name ───────────────────────────
COLUMN_MAP = {
    'First Name':                 'first_name',
    'Last Name':                  'last_name',
    'Date of Birth':              'date_of_birth',
    'Gender':                     'gender',
    'Phone Number':               'phone_number',
    'Email':                      'email',
    'Address':                    'address',
    'Blood Group':                'blood_group',
    'Genotype':                   'genotype',
    'Allergies':                  'allergies',
    'Chronic Conditions':         'chronic_conditions',
    'Next of Kin Name':           'nok_name',
    'Next of Kin Phone':          'nok_phone',
    'Next of Kin Relationship':   'nok_relationship',
    'Old Hospital Number':        'legacy_hospital_number',
}

REQUIRED_FIELDS = ['first_name', 'last_name', 'date_of_birth', 'gender', 'phone_number']
OPTIONAL_FIELDS = [
    'email', 'address', 'blood_group', 'genotype',
    'allergies', 'chronic_conditions',
    'nok_name', 'nok_phone', 'nok_relationship',
    'legacy_hospital_number',
]

GENDER_MAP   = {'male':'M','m':'M','female':'F','f':'F','other':'O','o':'O'}
BLOOD_GROUPS = ['A+','A-','B+','B-','AB+','AB-','O+','O-']
GENOTYPES    = ['AA','AS','SS','SC','AC']


def _next_hospital_number(year):
    """Auto-generate the next ABI hospital number for the given year."""
    from apps.patients.models import Patient
    last = Patient.objects.filter(
        hospital_number__startswith=f'ABI-{year}-'
    ).order_by('-hospital_number').first()
    seq = int(last.hospital_number.split('-')[-1]) + 1 if last else 1
    return f'ABI-{year}-{str(seq).zfill(5)}'


# ─────────────────────────────────────────────────────────────────────
# IMPORT
# ─────────────────────────────────────────────────────────────────────

@login_required
def import_patients_view(request):
    if not (request.user.is_admin_staff or request.user.is_receptionist):
        messages.error(request, "Permission denied.")
        return redirect('patients:dashboard')

    if request.method == 'POST':
        uploaded_file = request.FILES.get('excel_file')

        if not uploaded_file:
            messages.error(request, "Please select an Excel file.")
            return redirect('imports:import_patients')

        if not uploaded_file.name.lower().endswith(('.xlsx', '.xls')):
            messages.error(request, "Only .xlsx and .xls files are accepted.")
            return redirect('imports:import_patients')

        session = ImportSession.objects.create(
            uploaded_by=request.user,
            file_name=uploaded_file.name,
            status='processing',
        )

        try:
            _run_import(session, uploaded_file, request)
        except Exception as exc:
            session.status = 'failed'
            session.notes  = str(exc)
            session.save()
            messages.error(request, f"Import failed: {exc}")
            return redirect('imports:import_patients')

        log_action(request.user, 'CREATE', request,
                   f"Excel import: {session.success_count} patients added "
                   f"from {uploaded_file.name}")

        if session.error_count > 0:
            messages.warning(request,
                f"Import complete — "
                f"{session.success_count} patients added, "
                f"{session.error_count} rows flagged for review, "
                f"{session.skip_count} duplicates skipped.")
        else:
            messages.success(request,
                f"Import complete — {session.success_count} patients added successfully.")

        return redirect('imports:session_detail', pk=session.pk)

    recent = ImportSession.objects.all()[:10]
    return render(request, 'imports/import_patients.html', {
        'page_title':      'Import Patients from Excel',
        'required_fields': REQUIRED_FIELDS,
        'optional_fields': OPTIONAL_FIELDS,
        'recent_sessions': recent,
        'column_map':      COLUMN_MAP,
    })


def _run_import(session, uploaded_file, request):
    """Process every data row in the Excel file."""
    try:
        import openpyxl
    except ImportError:
        raise ImportError(
            "openpyxl is not installed. Run: pip install openpyxl"
        )

    from apps.patients.models import Patient

    wb   = openpyxl.load_workbook(uploaded_file, data_only=True)
    ws   = wb.active
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        raise ValueError("The Excel file is empty.")

    # Map header text → column index
    raw_headers = [str(c).strip() if c else '' for c in rows[0]]
    col_idx = {}
    for i, header in enumerate(raw_headers):
        field = COLUMN_MAP.get(header)
        if field:
            col_idx[field] = i

    data_rows = rows[1:]
    session.total_rows = len(data_rows)
    session.save()

    year = timezone.now().year

    for row_num, row in enumerate(data_rows, start=2):

        # Skip blank rows
        if all(c is None or str(c).strip() == '' for c in row):
            session.skip_count += 1
            continue

        # Pull values out of the row
        row_data = {}
        for field, idx in col_idx.items():
            row_data[field] = str(row[idx]).strip() if idx < len(row) and row[idx] is not None else ''

        # ── Required field check ────────────────────────────────────
        missing = [f for f in REQUIRED_FIELDS if not row_data.get(f)]
        if missing:
            ImportError.objects.create(
                session=session,
                row_number=row_num,
                error_type='missing_required',
                field_name=', '.join(missing),
                error_message=(
                    f"Missing required field(s): "
                    f"{', '.join(f.replace('_',' ').title() for f in missing)}"
                ),
                raw_data=row_data,
            )
            session.error_count += 1
            continue

        # ── Gender ──────────────────────────────────────────────────
        gender = GENDER_MAP.get(row_data.get('gender', '').lower())
        if not gender:
            ImportError.objects.create(
                session=session, row_number=row_num,
                error_type='invalid_format', field_name='gender',
                error_message=(
                    f"Invalid gender value: '{row_data.get('gender')}'. "
                    f"Use Male, Female, or Other."
                ),
                raw_data=row_data,
            )
            session.error_count += 1
            continue

        # ── Blood group ──────────────────────────────────────────────
        blood_group = row_data.get('blood_group', '').strip().upper()
        if blood_group and blood_group not in BLOOD_GROUPS:
            ImportError.objects.create(
                session=session, row_number=row_num,
                error_type='invalid_format', field_name='blood_group',
                error_message=(
                    f"Invalid blood group: '{blood_group}'. "
                    f"Use one of: {', '.join(BLOOD_GROUPS)}."
                ),
                raw_data=row_data,
            )
            session.error_count += 1
            continue

        # ── Date of birth ────────────────────────────────────────────
        import datetime
        dob     = None
        dob_raw = row_data.get('date_of_birth', '')
        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%m/%d/%Y'):
            try:
                dob = datetime.datetime.strptime(dob_raw, fmt).date()
                break
            except ValueError:
                continue

        if dob is None:
            ImportError.objects.create(
                session=session, row_number=row_num,
                error_type='invalid_format', field_name='date_of_birth',
                error_message=(
                    f"Cannot read date: '{dob_raw}'. "
                    f"Use YYYY-MM-DD or DD/MM/YYYY."
                ),
                raw_data=row_data,
            )
            session.error_count += 1
            continue

        # ── Duplicate check (name + DOB) ─────────────────────────────
        existing = Patient.objects.filter(
            first_name__iexact=row_data['first_name'],
            last_name__iexact=row_data['last_name'],
            date_of_birth=dob,
            is_active=True,
        ).first()

        if existing:
            ImportError.objects.create(
                session=session, row_number=row_num,
                error_type='duplicate', field_name='',
                error_message=(
                    f"Patient {row_data['first_name']} {row_data['last_name']} "
                    f"(DOB {dob}) already exists as {existing.hospital_number}. "
                    f"Flagged for admin review."
                ),
                raw_data=row_data,
            )
            session.skip_count += 1
            continue

        # ── Create patient ────────────────────────────────────────────
        hospital_number = _next_hospital_number(year)
        genotype = row_data.get('genotype', '').strip().upper()
        if genotype not in GENOTYPES:
            genotype = ''

        try:
            patient = Patient(
                hospital_number=hospital_number,
                first_name=row_data['first_name'],
                last_name=row_data['last_name'],
                date_of_birth=dob,
                gender=gender,
                phone_number=row_data.get('phone_number', ''),
                email=row_data.get('email', ''),
                address=row_data.get('address', ''),
                blood_group=blood_group,
                genotype=genotype,
                allergies=row_data.get('allergies', ''),
                chronic_conditions=row_data.get('chronic_conditions', ''),
                nok_name=row_data.get('nok_name', ''),
                nok_phone=row_data.get('nok_phone', ''),
                nok_relationship=row_data.get('nok_relationship', ''),
                legacy_hospital_number=row_data.get('legacy_hospital_number', ''),
                registered_by=request.user,
                is_active=True,
            )
            patient.full_clean()
            patient.save()
            session.success_count += 1
        except Exception as exc:
            ImportError.objects.create(
                session=session, row_number=row_num,
                error_type='other', field_name='',
                error_message=str(exc),
                raw_data=row_data,
            )
            session.error_count += 1

    session.status = (
        'complete_with_errors' if session.error_count > 0 else 'complete'
    )
    session.save()


@login_required
def session_detail_view(request, pk):
    """Show the result of one import session including all flagged rows."""
    session  = get_object_or_404(ImportSession, pk=pk)
    errors   = ImportError.objects.filter(session=session, is_resolved=False)
    resolved = ImportError.objects.filter(session=session, is_resolved=True)

    return render(request, 'imports/session_detail.html', {
        'page_title': f"Import #{pk} — {session.file_name}",
        'session':    session,
        'errors':     errors,
        'resolved':   resolved,
    })


# ─────────────────────────────────────────────────────────────────────
# EXPORT
# ─────────────────────────────────────────────────────────────────────

@login_required
def export_patients_view(request):
    """Export all active patients to a formatted Excel file."""
    if not (request.user.is_admin_staff or request.user.is_receptionist):
        messages.error(request, "Permission denied.")
        return redirect('patients:dashboard')

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        messages.error(request, "openpyxl is not installed. Run: pip install openpyxl")
        return redirect('patients:dashboard')

    from apps.patients.models import Patient

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Patients"

    # ── Styling ──────────────────────────────────────────────────────
    header_font  = Font(bold=True, color='FFFFFF', size=11)
    header_fill  = PatternFill(
        start_color='0A5C8A', end_color='0A5C8A', fill_type='solid'
    )
    header_align = Alignment(horizontal='center', vertical='center')

    headers = [
        'Hospital Number', 'Old Hospital Number',
        'First Name', 'Last Name',
        'Date of Birth', 'Age', 'Gender',
        'Phone Number', 'Email',
        'Blood Group', 'Genotype',
        'Allergies', 'Chronic Conditions',
        'Address',
        'Next of Kin Name', 'Next of Kin Phone', 'Next of Kin Relationship',
        'Date Registered', 'Status',
    ]

    for col_num, header in enumerate(headers, 1):
        cell             = ws.cell(row=1, column=col_num, value=header)
        cell.font        = header_font
        cell.fill        = header_fill
        cell.alignment   = header_align

    ws.row_dimensions[1].height = 22

    # ── Data rows ─────────────────────────────────────────────────────
    patients = Patient.objects.filter(is_active=True).order_by('hospital_number')

    for row_num, p in enumerate(patients, start=2):
        ws.cell(row=row_num, column=1,  value=p.hospital_number)
        ws.cell(row=row_num, column=2,  value=p.legacy_hospital_number or '')
        ws.cell(row=row_num, column=3,  value=p.first_name)
        ws.cell(row=row_num, column=4,  value=p.last_name)
        ws.cell(row=row_num, column=5,  value=str(p.date_of_birth))
        ws.cell(row=row_num, column=6,  value=p.age)
        ws.cell(row=row_num, column=7,  value=p.get_gender_display())
        ws.cell(row=row_num, column=8,  value=p.phone_number)
        ws.cell(row=row_num, column=9,  value=p.email)
        ws.cell(row=row_num, column=10, value=p.blood_group)
        ws.cell(row=row_num, column=11, value=p.genotype)
        ws.cell(row=row_num, column=12, value=p.allergies)
        ws.cell(row=row_num, column=13, value=p.chronic_conditions)
        ws.cell(row=row_num, column=14, value=p.address)
        ws.cell(row=row_num, column=15, value=p.nok_name)
        ws.cell(row=row_num, column=16, value=p.nok_phone)
        ws.cell(row=row_num, column=17, value=p.nok_relationship)
        ws.cell(row=row_num, column=18, value=str(p.created_at.date()))
        ws.cell(row=row_num, column=19, value='Active')

    # ── Auto-fit columns ──────────────────────────────────────────────
    for col in ws.columns:
        max_len = max(
            (len(str(cell.value or '')) for cell in col), default=10
        )
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 42)

    # ── Second sheet: blank import template ───────────────────────────
    ws2 = wb.create_sheet("Import Template")
    import_headers = list(COLUMN_MAP.keys())
    for col_num, header in enumerate(import_headers, 1):
        cell           = ws2.cell(row=1, column=col_num, value=header)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align

    # Example row so users know the format
    examples = {
        'First Name': 'John', 'Last Name': 'Olusun',
        'Date of Birth': '1990-01-15', 'Gender': 'Male',
        'Phone Number': '08012345678', 'Email': 'john@email.com',
        'Blood Group': 'O+', 'Genotype': 'AA',
    }
    for col_num, header in enumerate(import_headers, 1):
        ws2.cell(row=2, column=col_num, value=examples.get(header, ''))

    for col in ws2.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=10)
        ws2.column_dimensions[col[0].column_letter].width = min(max_len + 4, 36)

    # ── Build response ────────────────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"abicare_patients_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"

    log_action(request.user, 'VIEW', request,
               f"Exported {patients.count()} patients to Excel")

    response = HttpResponse(
        buffer.getvalue(),
        content_type=(
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        ),
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def download_template_view(request):
    """Download a blank Excel import template (no patient data)."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        messages.error(request, "openpyxl is not installed. Run: pip install openpyxl")
        return redirect('imports:import_patients')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Patient Import Template"

    header_font  = Font(bold=True, color='FFFFFF')
    header_fill  = PatternFill(
        start_color='0A5C8A', end_color='0A5C8A', fill_type='solid'
    )

    headers = list(COLUMN_MAP.keys())
    for col_num, header in enumerate(headers, 1):
        cell        = ws.cell(row=1, column=col_num, value=header)
        cell.font   = header_font
        cell.fill   = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Example row
    examples = {
        'First Name': 'John', 'Last Name': 'Olusun',
        'Date of Birth': '1990-01-15', 'Gender': 'Male',
        'Phone Number': '08012345678', 'Email': 'john@email.com',
        'Blood Group': 'O+', 'Genotype': 'AA',
        'Next of Kin Name': 'Mary Olusun',
        'Next of Kin Phone': '08098765432',
        'Next of Kin Relationship': 'Spouse',
    }
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=2, column=col_num, value=examples.get(header, ''))

    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 36)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type=(
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        ),
    )
    response['Content-Disposition'] = (
        'attachment; filename="abicare_patient_import_template.xlsx"'
    )
    return response

# Create your views here.
