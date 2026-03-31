"""
AbiCare - Excel Import / Export
================================
Rules:
1. Hospital numbers always auto-generated — never from Excel
2. Full Name column → first word = first_name, rest = last_name
3. Missing required fields filled with PENDING + flagged red in preview
4. Family detection:
   - filename contains "family" → all rows grouped into one family named after the file
   - "Family Group" column → rows with same value grouped into that family
   - Both rules can apply simultaneously
5. Errors are saved, never block the import
"""

import io
import os
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from django.utils import timezone

from .models import ImportSession, ImportError
from apps.audit_logs.utils import log_action


COLUMN_MAP = {
    'Full Name':                  'full_name',       # split to first+last
    'First Name':                 'first_name',
    'Last Name':                  'last_name',
    'Date of Birth':              'date_of_birth',
    'Gender':                     'gender',
    'Phone Number':               'phone_number',
    'Email':                      'email',
    'Address':                    'address',
    'City':                       'city',
    'State':                      'state',
    'LGA':                        'lga',
    'Hometown':                   'hometown',
    'State of Origin':            'state_of_origin',
    'Nationality':                'nationality',
    'Religion':                   'religion',
    'Occupation':                 'occupation',
    'Marital Status':             'marital_status',
    'Blood Group':                'blood_group',
    'Genotype':                   'genotype',
    'Allergies':                  'allergies',
    'Chronic Conditions':         'chronic_conditions',
    'Next of Kin Name':           'nok_name',
    'Next of Kin Phone':          'nok_phone',
    'Next of Kin Relationship':   'nok_relationship',
    'Insurance Provider':         'insurance_provider',
    'Insurance Number':           'insurance_number',
    'NHIS Number':                'nhis_number',
    'Old Hospital Number':        'legacy_hospital_number',
    'Family Group':               'family_group',    # family grouping
}

REQUIRED_FIELDS = ['first_name', 'last_name', 'date_of_birth', 'gender', 'phone_number']
GENDER_MAP   = {'male':'M','m':'M','man':'M','boy':'M','female':'F','f':'F','woman':'F','girl':'F','other':'O','o':'O'}
BLOOD_GROUPS = ['A+','A-','B+','B-','AB+','AB-','O+','O-']
GENOTYPES    = ['AA','AS','SS','SC','AC']
PENDING      = 'PENDING'


def _next_number(year):
    from apps.patients.models import Patient
    last = Patient.objects.filter(
        hospital_number__startswith=f'ABI-{year}-'
    ).order_by('-hospital_number').first()
    seq = int(last.hospital_number.split('-')[-1]) + 1 if last else 1
    return f'ABI-{year}-{str(seq).zfill(5)}'


def _split_full_name(full_name):
    """'John Olusun Adeyemi' → first='John', last='Olusun Adeyemi'"""
    parts = full_name.strip().split()
    if not parts:
        return PENDING, PENDING
    first = parts[0]
    last  = ' '.join(parts[1:]) if len(parts) > 1 else PENDING
    return first, last


def _get_or_create_family(family_name, created_by):
    from apps.families.models import FamilyGroup
    family, _ = FamilyGroup.objects.get_or_create(
        family_name__iexact=family_name,
        defaults={'family_name': family_name, 'created_by': created_by, 'is_active': True}
    )
    return family


def _add_to_family(patient, family, relationship='other', added_by=None):
    from apps.families.models import FamilyMember
    FamilyMember.objects.get_or_create(
        family=family, patient=patient,
        defaults={'relationship': relationship, 'added_by': added_by}
    )


@login_required
def import_patients_view(request):
    if not (request.user.is_admin_staff or request.user.is_receptionist):
        messages.error(request, "Permission denied.")
        return redirect('patients:dashboard')

    if request.method == 'POST':
        uploaded = request.FILES.get('excel_file')
        if not uploaded:
            messages.error(request, "Please select an Excel file.")
            return redirect('imports:import_patients')
        if not uploaded.name.lower().endswith(('.xlsx', '.xls')):
            messages.error(request, "Only .xlsx and .xls files are accepted.")
            return redirect('imports:import_patients')

        session = ImportSession.objects.create(
            uploaded_by=request.user,
            file_name=uploaded.name,
            status='processing',
        )
        try:
            _run_import(session, uploaded, request)
        except Exception as e:
            session.status = 'failed'
            session.notes  = str(e)
            session.save()
            messages.error(request, f"Import failed: {e}")
            return redirect('imports:import_patients')

        log_action(request.user, 'CREATE', request,
                   f"Imported {session.success_count} patients from {uploaded.name}")

        if session.error_count > 0:
            messages.warning(request,
                f"Import done — {session.success_count} patients added, "
                f"{session.error_count} rows had errors (flagged for review), "
                f"{session.skip_count} duplicates skipped.")
        else:
            messages.success(request,
                f"Import complete — {session.success_count} patients added.")
        return redirect('imports:session_detail', pk=session.pk)

    recent = ImportSession.objects.all()[:10]
    return render(request, 'imports/import_patients.html', {
        'page_title':      'Import Patients from Excel',
        'required_fields': REQUIRED_FIELDS,
        'column_map':      COLUMN_MAP,
        'recent_sessions': recent,
    })


def _run_import(session, uploaded_file, request):
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is not installed. Run: pip install openpyxl")

    from apps.patients.models import Patient

    wb    = openpyxl.load_workbook(uploaded_file, data_only=True)
    ws    = wb.active
    rows  = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("The Excel file is empty.")

    raw_headers = [str(c).strip() if c else '' for c in rows[0]]
    col_idx = {}
    for i, h in enumerate(raw_headers):
        field = COLUMN_MAP.get(h)
        if field:
            col_idx[field] = i

    data_rows = rows[1:]
    session.total_rows = len(data_rows)
    session.save()

    year = timezone.now().year

    # ── Family from filename ──────────────────────────────────────────
    filename_lower  = uploaded_file.name.lower()
    filename_family = None
    if 'family' in filename_lower:
        # Use the filename (without extension) as family name
        base = os.path.splitext(uploaded_file.name)[0]
        # Capitalise nicely e.g. "olusun_family" → "Olusun Family"
        filename_family = base.replace('_', ' ').replace('-', ' ').title()

    for row_num, row in enumerate(data_rows, start=2):
        if all(c is None or str(c).strip() == '' for c in row):
            session.skip_count += 1
            continue

        # Pull raw values
        rv = {}
        for field, idx in col_idx.items():
            rv[field] = str(row[idx]).strip() if idx < len(row) and row[idx] is not None else ''

        # ── Handle Full Name → first + last ─────────────────────────
        if 'full_name' in rv and rv['full_name']:
            fn, ln = _split_full_name(rv['full_name'])
            rv.setdefault('first_name', fn)
            rv.setdefault('last_name',  ln)
        rv.pop('full_name', None)

        # ── Detect family column ─────────────────────────────────────
        row_family_name = rv.pop('family_group', '').strip()

        # ── Fill missing required fields with PENDING ────────────────
        pending_fields = []
        for field in REQUIRED_FIELDS:
            if not rv.get(field):
                rv[field] = PENDING
                pending_fields.append(field)

        # ── Gender normalise ─────────────────────────────────────────
        gender_raw = rv.get('gender', '').lower()
        gender     = GENDER_MAP.get(gender_raw)
        if not gender:
            if rv['gender'] == PENDING:
                gender = 'O'  # default — flagged anyway
            else:
                ImportError.objects.create(
                    session=session, row_number=row_num,
                    error_type='invalid_format', field_name='gender',
                    error_message=f"Invalid gender: '{rv['gender']}'. Use Male/Female/Other.",
                    raw_data=rv,
                )
                session.error_count += 1
                continue

        # ── Date parse ───────────────────────────────────────────────
        import datetime
        dob = None
        dob_raw = rv.get('date_of_birth', '')
        if dob_raw != PENDING:
            for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%m/%d/%Y'):
                try:
                    dob = datetime.datetime.strptime(dob_raw, fmt).date()
                    break
                except ValueError:
                    continue
            if dob is None:
                ImportError.objects.create(
                    session=session, row_number=row_num,
                    error_type='invalid_format', field_name='date_of_birth',
                    error_message=f"Cannot read date: '{dob_raw}'. Use YYYY-MM-DD.",
                    raw_data=rv,
                )
                session.error_count += 1
                continue
        else:
            # Use a placeholder DOB
            dob = datetime.date(1900, 1, 1)

        # ── Duplicate check ──────────────────────────────────────────
        if rv['first_name'] != PENDING and rv['last_name'] != PENDING:
            existing = Patient.objects.filter(
                first_name__iexact=rv['first_name'],
                last_name__iexact=rv['last_name'],
                date_of_birth=dob,
                is_active=True,
            ).first()
            if existing:
                ImportError.objects.create(
                    session=session, row_number=row_num,
                    error_type='duplicate', field_name='',
                    error_message=(
                        f"{rv['first_name']} {rv['last_name']} (DOB {dob}) "
                        f"already exists as {existing.hospital_number}."
                    ),
                    raw_data=rv,
                )
                session.skip_count += 1
                continue

        # ── Blood group / genotype ───────────────────────────────────
        blood_group = rv.get('blood_group', '').strip().upper()
        if blood_group and blood_group not in BLOOD_GROUPS:
            blood_group = ''
        genotype = rv.get('genotype', '').strip().upper()
        if genotype not in GENOTYPES:
            genotype = ''

        # ── Normalise choice fields to lowercase ─────────────────────
        marital_raw = rv.get('marital_status', '').strip().lower()
        valid_marital = ['single','married','divorced','widowed','separated']
        marital_status = marital_raw if marital_raw in valid_marital else ''

        religion_raw = rv.get('religion', '').strip().lower()
        valid_religion = ['christianity','islam','traditional','other','none']
        religion = religion_raw if religion_raw in valid_religion else ''

        # ── Create patient ───────────────────────────────────────────
        hospital_number = _next_number(year)
        has_pending = bool(pending_fields)

        try:
            patient = Patient(
                hospital_number=hospital_number,
                first_name=rv.get('first_name', PENDING),
                last_name=rv.get('last_name', PENDING),
                date_of_birth=dob,
                gender=gender,
                phone_number=rv.get('phone_number', PENDING),
                email=rv.get('email', ''),
                address=rv.get('address', ''),
                city=rv.get('city', ''),
                state=rv.get('state', ''),
                lga=rv.get('lga', ''),
                hometown=rv.get('hometown', ''),
                state_of_origin=rv.get('state_of_origin', ''),
                nationality=rv.get('nationality', 'Nigerian'),
                religion=religion,
                occupation=rv.get('occupation', ''),
                marital_status=marital_status,
                blood_group=blood_group,
                genotype=genotype,
                allergies=rv.get('allergies', ''),
                chronic_conditions=rv.get('chronic_conditions', ''),
                nok_name=rv.get('nok_name', ''),
                nok_phone=rv.get('nok_phone', ''),
                nok_relationship=rv.get('nok_relationship', ''),
                insurance_provider=rv.get('insurance_provider', ''),
                insurance_number=rv.get('insurance_number', ''),
                nhis_number=rv.get('nhis_number', ''),
                legacy_hospital_number=rv.get('legacy_hospital_number', ''),
                registered_by=request.user,
                is_active=True,
                has_pending_fields=has_pending,
                pending_field_list=','.join(pending_fields),
            )
            patient.save()
            session.success_count += 1

            # ── Family assignment ─────────────────────────────────────
            if filename_family:
                fam = _get_or_create_family(filename_family, request.user)
                _add_to_family(patient, fam, added_by=request.user)

            if row_family_name:
                fam2 = _get_or_create_family(row_family_name, request.user)
                _add_to_family(patient, fam2, added_by=request.user)

        except Exception as e:
            ImportError.objects.create(
                session=session, row_number=row_num,
                error_type='other', field_name='',
                error_message=str(e),
                raw_data=rv,
            )
            session.error_count += 1

    session.status = 'complete_with_errors' if session.error_count > 0 else 'complete'
    session.save()


@login_required
def session_detail_view(request, pk):
    session  = get_object_or_404(ImportSession, pk=pk)
    errors   = ImportError.objects.filter(session=session, is_resolved=False)
    resolved = ImportError.objects.filter(session=session, is_resolved=True)
    return render(request, 'imports/session_detail.html', {
        'page_title': f"Import #{pk} — {session.file_name}",
        'session':    session,
        'errors':     errors,
        'resolved':   resolved,
    })


@login_required
def export_patients_view(request):
    if not (request.user.is_admin_staff or request.user.is_receptionist):
        messages.error(request, "Permission denied.")
        return redirect('patients:dashboard')

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, PatternFill
        from openpyxl.styles.fills import PatternFill
    except ImportError:
        messages.error(request, "openpyxl is not installed. Run: pip install openpyxl")
        return redirect('patients:dashboard')

    from apps.patients.models import Patient

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "Patients"

    header_font  = Font(bold=True, color='FFFFFF', size=11)
    header_fill  = PatternFill(start_color='0A5C8A', end_color='0A5C8A', fill_type='solid')
    pending_fill = PatternFill(start_color='FED7D7', end_color='FED7D7', fill_type='solid')
    pending_font = Font(color='C53030', bold=True)
    header_align = Alignment(horizontal='center', vertical='center')

    headers = [
        'Hospital Number','Old Number','First Name','Last Name','Date of Birth',
        'Age','Gender','Phone','Alt Phone','Email','Marital Status','Religion',
        'Occupation','Address','City','State','LGA','Hometown','State of Origin',
        'Nationality','Blood Group','Genotype','Allergies','Chronic Conditions',
        'NOK Name','NOK Phone','NOK Relationship','Insurance','NHIS',
        'Date Registered','Status','Pending Fields',
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font; cell.fill = header_fill; cell.alignment = header_align
    ws.row_dimensions[1].height = 22

    patients = Patient.objects.filter(is_active=True).order_by('hospital_number')

    for r, p in enumerate(patients, start=2):
        row_values = [
            p.hospital_number, p.legacy_hospital_number, p.first_name, p.last_name,
            str(p.date_of_birth), p.age, p.get_gender_display(), p.phone_number,
            p.alt_phone_number, p.email, p.marital_status, p.religion,
            p.occupation, p.address, p.city, p.state, p.lga, p.hometown,
            p.state_of_origin, p.nationality, p.blood_group, p.genotype,
            p.allergies, p.chronic_conditions,
            p.nok_name, p.nok_phone, p.nok_relationship,
            p.insurance_provider, p.nhis_number,
            str(p.created_at.date()), 'Active',
            p.pending_field_list if p.has_pending_fields else '',
        ]
        for col, val in enumerate(row_values, 1):
            cell = ws.cell(row=r, column=col, value=val or '')
            # Highlight PENDING values red
            if str(val or '').upper() == 'PENDING':
                cell.fill = pending_fill
                cell.font = pending_font

    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    # Template sheet
    ws2 = wb.create_sheet("Import Template")
    template_headers = list(COLUMN_MAP.keys())
    for col, h in enumerate(template_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font; cell.fill = header_fill
    example = {'Full Name':'John Olusun','Date of Birth':'1990-01-15','Gender':'Male',
                'Phone Number':'08012345678','Email':'john@email.com','Family Group':'Olusun Family'}
    for col, h in enumerate(template_headers, 1):
        ws2.cell(row=2, column=col, value=example.get(h, ''))
    for col in ws2.columns:
        ws2.column_dimensions[col[0].column_letter].width = 22

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)

    filename = f"abicare_patients_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
    log_action(request.user, 'VIEW', request, f"Exported {patients.count()} patients")
    response = HttpResponse(buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def download_template_view(request):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        messages.error(request, "openpyxl not installed.")
        return redirect('imports:import_patients')

    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Import Template"
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='0A5C8A', end_color='0A5C8A', fill_type='solid')
    headers = list(COLUMN_MAP.keys())
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font; c.fill = header_fill
        c.alignment = Alignment(horizontal='center')
    example = {'Full Name':'John Olusun Adeyemi','Date of Birth':'1990-01-15',
               'Gender':'Male','Phone Number':'08012345678','Email':'john@email.com',
               'Blood Group':'O+','Genotype':'AA','Family Group':'Olusun Family',
               'State of Origin':'Oyo','Nationality':'Nigerian'}
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=example.get(h, ''))
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 22

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    response = HttpResponse(buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="abicare_import_template.xlsx"'
    return response
# Create your views here.
